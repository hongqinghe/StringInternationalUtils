# 将 help xml 文件转为带键值key-value的形式
from openpyxl.workbook import Workbook
# 颜色
from openpyxl.styles import  PatternFill
from openpyxl.styles import Font, Border,Side,Alignment

from help_string import help_xml_constants as HC, help_xml_parse_utils as H_UTILS

#  步骤一


# <WarehouseManagerViewHelp version="V1" name="仓库管理">
#     <![CDATA[
# <b>添加仓库</b>
#     ▪︎︎点添加按钮，可添加新的仓库。

# <b>供货仓库</b>
#     ▪︎︎仓库可设置为供货仓库。
#     ▪︎︎设置为供货仓库的仓库，门店在创建采购单和退货单时，可指定向该仓库进行采购或退货。
#     ▪︎︎总部可设置多个供货仓库，也可不设置任何供货仓库。不设置供货仓库时，门店仍可向总部采购，只是不能指定向哪个仓库采购。
#     ]]>
# </WarehouseManagerViewHelp

# excel 输出结构
#  key                              value                               en
#  key_version                      title
#  WarehouseManagerViewHelp_V1      仓库管理
#  key_A_0                          添加仓库（截取掉<b>*</b>）
#  key_A_1                          点添加按钮，可添加新的仓库。(截取掉    ▪︎︎)
#  key_B_0                          供货仓库（截取掉<b>*</b>）
#  key_B_1                          仓库可设置为供货仓库。(截取掉    ▪︎︎)
#  ...依次递加

#  en 在excel文件的位置  行
_en_sheet_location=1
# thai 在 excel文件的位置  行
_thai_sheet_loaction=1

errorList=[]

xmlList=[]
# worksheets
ws=None
# 是否开启容错处理 可以在生成的文件当中进行标记那个输出不对应
isOpenFaultTolerantProcess=True

test_error_path = r'/Users/hehongqing/Downloads/error_list.xml'
sheet_title = 'help'

# 写入表格当中 并且方法内部提供给两种提示  容错【isOpenFaultTolerantProcess 开关】、 未翻译标记
def writeWs(xmlKey, xmlValue, sheetRowTitle, location,errorColor,noTranslateColor):
    sheetRowErrowTitle = HC._sheet_title_error_location
    # 进行判断
    cell = ws[HC._sheet_title_key_location+str(location)]

    zhKey=cell._value
    # 如果这里的key无法进行匹配 直接标红 需要检查xml文件
    if not zhKey == xmlKey and isOpenFaultTolerantProcess:
        cell = ws[sheetRowErrowTitle+str(location)]
        cell.fill = PatternFill(
            fill_type='lightGray', bgColor=errorColor, fgColor=errorColor)
        ws[sheetRowErrowTitle+str(location)] = zhKey
        # print('出错key打印如下    cellValue========',havekey,'xmlKey============',xmlKey)
        # errorList.append('出错key打印如下    cellValue======'+havekey+'======xmlKey========'+xmlKey)
    ws[HC._sheet_title_key_location+str(location)] = xmlKey

    # 获取中文cell
    cell = ws[HC._sheet_title_zh_location+str(location)]
    # 中文一致，则认为没有进行翻译，输出值为空，并且做出标记
    if not cell._value=='' and cell._value == xmlValue:
        cell=ws[sheetRowTitle+str(location)]
        cell.fill = PatternFill(
            fill_type='lightGray', bgColor=noTranslateColor, fgColor=noTranslateColor)
        ws[sheetRowTitle+str(location)]=''
    else:
        ws[sheetRowTitle+str(location)] = xmlValue

# 多语言处理 处理除了中文以外的国际化语言格式
def multipleTypeToDeal(xmlKey, xmlValue, type):
    global _en_sheet_location, _thai_sheet_loaction
    sheetRowTitle=None
    location=1
    errorColor = 'e3b8e5'
    noTranslateColor=''
    if type == HC._FILE_TYPE_EN:
        _en_sheet_location += 1
        location=_en_sheet_location
        sheetRowTitle = HC._sheet_title_en_location
        noTranslateColor = 'dcbdea'
    elif type == HC._FILE_TYPE_THAI:
        _thai_sheet_loaction+=1
        location = _thai_sheet_loaction
        sheetRowTitle = HC._sheet_title_thai_location
        noTranslateColor = 'bae0e5'
    writeWs(xmlKey, xmlValue, sheetRowTitle, location,errorColor, noTranslateColor)

def generateExcel(xmlKey,xmlValue,type):
    # global _en_sheet_location, _thai_sheet_loaction
    if xmlKey=='':
        return
    # 根据类型输出到excel中 zh 直接 append其他语言根据之前设定的位置进行一次递加
    if type == HC._FILE_TYPE_ZH:
        # 在这里输出title到excel中  excel_key_title--->nodeTitleValue
        ws.append({1: xmlKey, 2: xmlValue})
    else:
        multipleTypeToDeal(xmlKey,xmlValue,type)


def readFile(filePath, type):
    file = open(filePath)
    while 1:
        line = file.readline()
        if not line:
            file.close
            break
        else:
            # 读取到的每行
            # print(line)
            xml_key, xml_value = H_UTILS.dealLine(line)
            generateExcel(xml_key, xml_value, type)

def initExcel():
    wb = Workbook()
    return wb
def getSheetWs(wb,sheetTitle):
    ws = wb.worksheets[0]
    ws.title = sheetTitle
    ws.append({1: 'Key', 2: 'Value', 3: 'En', 4: 'Thai'})


    currentRow=ws._current_row
    cellHorizontal='center'

    cell=ws['A'+str(currentRow)]
    aCellColor='d3d3d3'
    cell.alignment=Alignment(horizontal=cellHorizontal,vertical=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=aCellColor,fgColor=aCellColor)

    cell=ws['B'+str(currentRow)] 
    bCellColor='fff0f5'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=bCellColor,fgColor=bCellColor)

    cell=ws['C'+str(currentRow)]
    cCellColor='ffe4e1'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=cCellColor,fgColor=cCellColor)

    cell=ws['D'+str(currentRow)]
    cell.alignment=Alignment(horizontal=cellHorizontal)
    dCellColor='ffc0cb'
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=dCellColor,fgColor=dCellColor)

    return ws

# 设置每个单元格的默认属性
def setCellProperty(worksheet):
    for i in range(1, worksheet.max_row+1):
        # worksheet.row_dimensions[i].height=50
        for col in range(ord('A'),ord('F')):
            worksheet[chr(col)+str(i)].font=Font(size=16)
            worksheet[chr(col)+str(i)].alignment=Alignment(wrap_text=True)
            borderStype='thick'
            borderColor='556b2f'
            worksheet[chr(col)+str(i)].border=Border(left=Side(border_style=borderStype,color=borderColor),
            right=Side(border_style=borderStype,color=borderColor),
            top=Side(border_style=borderStype,color=borderColor),
            bottom=Side(border_style=borderStype,color=borderColor))

    worksheet.column_dimensions['A'].width=45
    worksheet.column_dimensions['B'].width=85
    worksheet.column_dimensions['C'].width=90
    worksheet.column_dimensions['D'].width=90
    worksheet.column_dimensions['E'].width=90

def xmlConversionExcel():
    wb = initExcel()
    global ws
    ws = getSheetWs(wb,sheet_title)
    
    # readFile(HC._TEMP_ZH_RESULT_PATH, HC. _FILE_TYPE_ZH)
    # readFile(HC._TEMP_EN_RESULT_PATH, HC._FILE_TYPE_EN)
    # readFile(HC._TEMP_THAI_RESULT_PATH, HC. _FILE_TYPE_THAI)

    readFile(HC._SOURCE_ZH_PATH, HC. _FILE_TYPE_ZH)
    readFile(HC._SOURCE_EN_PATH, HC._FILE_TYPE_EN)
    readFile(HC._SOURCE_THAI_PATH, HC. _FILE_TYPE_THAI)

    setCellProperty(ws)
    wb.save(HC._TO_TRANSLATE_EXCEL_PATH)

if __name__ == "__main__":
    xmlConversionExcel()

