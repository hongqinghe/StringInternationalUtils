# 将 help xml 文件转为带键值key-value的形式
import re
import os
from openpyxl.workbook import Workbook
# 颜色
from openpyxl.styles import  PatternFill
from openpyxl.styles import Color, Fill,Font,colors,Border,Side,Alignment
from openpyxl.cell import Cell

import help_xml_constants as HC


# <WarehouseManagerViewHelp version="V1" name="仓库管理">
#     <![CDATA[
# <b>添加仓库</b>
#     ▪︎︎点添加按钮，可添加新的仓库。

# <b>供货仓库</b>
#     ▪︎︎仓库可设置为供货仓库。
#     ▪︎︎设置为供货仓库的仓库，门店在创建采购单和退货单时，可指定向该仓库进行采购或退货。
#     ▪︎︎总部可设置多个供货仓库，也可不设置任何供货仓库。不设置供货仓库时，门店仍可向总部采购，只是不能指定向哪个仓库采购。
#     ]]>
# </WarehouseManagerViewHelp

# WarehouseManagerViewHelp
_help_key = None

# V1
_help_key_version = None

# 是否已经初始化key
_init_key = False

# 第一个节点拼接值 【A-Z】
# A-Z
_key_first_node = None

# 第二个节点拼接值 【1...】
# 1...
_key_second_node = None

# 第三个节点拼接值 【a-z】
# a-z
_key_third_node = None


# 第四个节点拼接值 【1...】
# 1...
_key_fourth_node = None


# 特殊节点拼接值 【1...】
# 1...
_special_node = None

# 特殊行进行特殊的标识
SPECIAL_FLAG = 'SP'

# 操作类型 
# 转化为excel 
_operator_type_excel = 'toExcel'

# 转化为 key-vaule 形式
_operator_type_key_xml = 'ToKeyXml'

#  en 在excel文件的位置  行
_en_sheet_location=1
# thai 在 excel文件的位置  行
_thai_sheet_loaction=1


errorList=[]

INVALID_STRING_1 = '    <![CDATA['
INVALID_STRING_2 = '    ]]>'
STRING_CONNECTOR = '_'

xmlList=[]
# worksheets
ws=None
# 是否开启容错处理 可以在生成的文件当中进行标记那个输出不对应
isOpenFaultTolerantProcess=True

help_zh_path = '/Users/hehongqing/workspace/utilsmaven/static_help/resources/xml/SupplyHelpFiles.xml'
# help_en_path = '/Users/hehongqing/WorkSpace/Android/help/static_help/resources/xml/SupplyHelpENFiles.xml'
help_en_path = '/Users/hehongqing/workspace/utilsmaven/static_help/resources/xml/SupplyHelpENFiles.xml'
help_thai_path = '/Users/hehongqing/Downloads/thai_help_string.xml'

# test_path = r'/Users/hehongqing/Downloads/test_help.xml'
test_path_key = r'/Users/hehongqing/Downloads/test_help_key.xml'
test_path_key_en = r'/Users/hehongqing/Downloads/test_help_key_en.xml'
test_path_key_thai= r'/Users/hehongqing/Downloads/test_help_key_thai.xml'
su_string_file_path = r'/Users/hehongqing/Downloads/help_string_3.xlsx'

test_error_path = r'/Users/hehongqing/Downloads/error_list.xml'
sheet_title = 'help'

# 写入表格当中 并且方法内部提供给两种提示  容错【isOpenFaultTolerantProcess 开关】、 未翻译标记
def writeWs(xmlKey, xmlValue, sheetRowTitle, location):
    sheetRowErrowTitle = HC._sheet_title_error_location
    # 进行判断
    cell = ws[HC._sheet_title_key_location+str(location)]
    havekey=cell._value
    # 如果这里的key无法进行匹配 直接标红 需要检查xml文件
    if not havekey == xmlKey and isOpenFaultTolerantProcess:
        cell = ws[sheetRowErrowTitle+str(location)]
        enEmptyColor = 'dc143c'
        cell.fill = PatternFill(
            fill_type='lightGray', bgColor=enEmptyColor, fgColor=enEmptyColor)
        ws[sheetRowErrowTitle+str(location)] = xmlKey
        print('出错key打印如下    cellValue========',havekey,'xmlKey============',xmlKey)
        errorList.append('出错key打印如下    cellValue======'+havekey+'======xmlKey========'+xmlKey)
    ws[HC._sheet_title_key_location+str(location)] = xmlKey

    cell = ws[HC._sheet_title_zh_location+str(location)]
    # 中文一致，则认为没有进行翻译，输出值为空，并且做出标记
    if cell._value == xmlValue:
        ws[sheetRowTitle+str(location)] = ''
        enEmptyColor = 'dda0dd'
        ws[sheetRowTitle+str(location)].fill == PatternFill(
            fill_type='lightGray', bgColor=enEmptyColor, fgColor=enEmptyColor)
    else:
        ws[sheetRowTitle+str(location)] = xmlValue

# 多语言处理 处理除了中文以外的国际化语言格式
def multipleTypeToDeal(xmlKey, xmlValue, type):
    global _en_sheet_location, _thai_sheet_loaction
    sheetRowTitle=None
    location=1
    if type == HC._FILE_TYPE_EN:
        _en_sheet_location += 1
        location=_en_sheet_location
        sheetRowTitle = HC._sheet_title_en_location
    elif type == HC._FILE_TYPE_THAI:
        _thai_sheet_loaction+=1
        location = _thai_sheet_loaction
        sheetRowTitle = HC._sheet_title_thai_location
    writeWs(xmlKey, xmlValue, sheetRowTitle,location)

def generateExcel(xmlKey,xmlValue,type):
    # global _en_sheet_location, _thai_sheet_loaction
    if xmlKey=='':
        return
    # 根据类型输出到excel中 zh 直接 append其他语言根据之前设定的位置进行一次递加
    if type == HC._FILE_TYPE_ZH:
        # 在这里输出title到excel中  excel_key_title--->nodeTitleValue
        ws.append({1: xmlKey, 2: xmlValue})
    else:
        multipleTypeToDeal(xmlKey,xmlValue,type)
    # elif type == FILE_TYPE_EN:
    #     _en_sheet_location += 1
    #     # 进行判断
    #     cell = ws['A'+str(_en_sheet_location)]
    #     # 如果这里的key无法进行匹配 直接标红 需要检查xml文件
    #     if not cell._value ==xmlKey and isOpenFaultTolerantProcess:
    #         cell = ws['E'+str(_en_sheet_location)]
    #         enEmptyColor='dc143c'
    #         cell.fill=PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
    #         ws['E'+str(_en_sheet_location)] = xmlKey

    #     ws['A'+str(_en_sheet_location)] = xmlKey
    #     cell=ws['B'+str(j)]
    #     # 和中文一致，则认为没有进行翻译，输出值为空，并且做出标记
    #     if cell._value==xmlValue:
    #         ws['C'+str(_en_sheet_location)] = ''
    #         enEmptyColor='dda0dd'
    #         ws['C'+str(_en_sheet_location)].fill==PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
    #     else:
    #         ws['C'+str(j)]=xmlValue
    # elif type == FILE_TYPE_THAI:
    #     _thai_sheet_loaction+=1
    #     # 进行判断
    #     cell = ws['A'+str(_thai_sheet_loaction)]
    #     # 如果这里的key无法进行匹配 直接标红 需要检查xml文件
    #     if not cell._value ==xmlKey and isOpenFaultTolerantProcess:
    #         cell = ws['E'+str(_thai_sheet_loaction)]
    #         enEmptyColor='dc143c'
    #         cell.fill=PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
    #         ws['E'+str(_thai_sheet_loaction)]=xmlKey

    #     ws['A'+str(_thai_sheet_loaction)]=xmlKey
    #     cell = ws['B'+str(_thai_sheet_loaction)]
    #     # 中文一致，则认为没有进行翻译，输出值为空，并且做出标记
    #     if cell._value==xmlValue:
    #         ws['D'+str(_thai_sheet_loaction)] = ''
    #         enEmptyColor='dda0dd'
    #         ws['D'+str(_thai_sheet_loaction)].fill == PatternFill(
    #             fill_type='lightGray', bgColor=enEmptyColor, fgColor=enEmptyColor)
    #     else:
    #         ws['D'+str(_thai_sheet_loaction)] = xmlValue


# 文件读取结束时进行文件的输出 ：格式  key---value
def  fileOperateFinish(type): 
    global xmlList
    if type == HC._FILE_TYPE_ZH:

        file=open(test_path_key,'w+')
    elif type == HC._FILE_TYPE_EN:
        file=open(test_path_key_en,'w+')
    elif type == HC._FILE_TYPE_THAI:
        file=open(test_path_key_thai,'w+')
    for  line in xmlList:
        file.write(line)
    file.close()

    xmlList.clear()


def readFile(filePath, type, operatorType):
    file = open(filePath)
    while 1:
        line = file.readline()
        if not line:
            file.close

            fileOperateFinish(type)
            break
        else:
            # 读取到的每行
            # print(line)

            xml_key,xml_value=dealLine(line, type)
            if operatorType == _operator_type_excel:
                generateExcel(xml_key,xml_value,type)
                appendKeyLine(xml_key, line)
            elif operatorType==_operator_type_key_xml:
                appendKeyLine(xml_key,line)

# 根据key---value 添加至 xmlList 当中
def  appendKeyLine(key,line):
    global xmlList
    # 如果这一行不含有\n 则认为这一行源数据输入不正确，需要校正(最后一行除外)
    if not '\n' in line and  not'</root>' in line:
            line=line+'\n'
    if key=='':
        xmlList.append(line)
    else: 
        xmlList.append('【key='+key+'】'+line)


def dealLine(line, type):
    global _init_key, _help_key, _key_first_node, _key_second_node, _key_third_node, _key_fourth_node, _special_node, _help_key_version
    if _init_key == False and _help_key == None:
        match = re.search(HC._help_node_title_version_pattern, line)
        #  pattern  <WarehouseManagerViewHelp version="V1" name="仓库管理">
        if not match == None:
            _init_key = True
            # print(match)
            nodeTitleMatch = re.search(HC._help_node_title_pattern, line)
            # print(nodeTitleMatch)
            # print('key          '+nodeTitleMatch.group(HC._help_node_title_pattern_key))

            _help_key = nodeTitleMatch.group(HC._help_node_title_pattern_key)

            nodeVersion = nodeTitleMatch.group(HC._help_node_title_pattern_version)
            # print('VersionName  '+nodeVersion)
            
            _help_key_version=nodeVersion

            nodeTitleValue = nodeTitleMatch.group(HC._help_node_title_pattern_title)
            # print('Title        '+nodeTitleValue)
            excel_key_title = _help_key+STRING_CONNECTOR+_help_key_version

            return  excel_key_title,nodeTitleValue
        else:
            return '',line
            # excel 输出结构
            #  key                              value                               en
            #  key_version                      title
            #  WarehouseManagerViewHelp_V1      仓库管理
            #  key_A_0                          添加仓库（截取掉<b>*</b>）
            #  key_A_1                          点添加按钮，可添加新的仓库。(截取掉    ▪︎︎)
            #  key_B_0                          供货仓库（截取掉<b>*</b>）
            #  key_B_1                          仓库可设置为供货仓库。(截取掉    ▪︎︎)
            #  ...依次递加
    elif INVALID_STRING_1 in line or INVALID_STRING_2 in line or line == '\n' or not re.search(r'^ *\n', line) == None or '<![CDATA[' in line or not re.search(r'	]]>',line)==None or ']]>'in line :
        # print('无效返回')
        return '',line
    elif not re.search(HC._help_node_b_pattern, line) == None:
        # pattern <b>添加仓库</b>
        nodeMatch = re.search(HC._help_node_b_pattern, line)
        # first_node_value = nodeMatch.group(HC._help_node_b_pattern_value)
        b_node_value = nodeMatch.group(HC._help_node_b_pattern_value)

        if _key_first_node == None:
            _key_first_node = 'A'
            # print(b_node_value)
            # 在这里输出 excel 第一个子节点  excel_key_node--->b_node_value
            excel_key_node = _help_key+_help_key_version+STRING_CONNECTOR+_key_first_node
            return  excel_key_node,b_node_value
        else:
            # A+1
            # 同时需要将节点 置为初始值
            _key_second_node = None
            _key_third_node = None
            _key_fourth_node = None

            _key_first_node = chr(ord(_key_first_node)+1)
            # 在这里输出 excel 第一个子节点  excel_key_node--->first_node_value
            excel_key_node = _help_key+_help_key_version+STRING_CONNECTOR+_key_first_node

            return excel_key_node,b_node_value

    elif not re.search(HC._help_node_first_pattern, line) == None:
        # pattern'    ▪︎︎请按照标准格式说出您想要采购的原料及其数量。'
        secondNodeMatch = re.search(HC._help_node_first_pattern, line)
        # print(secondNodeMatch)
        # second_node_value = secondNodeMatch.group(HC._help_node_first_pattern_value)
        first_node_value = secondNodeMatch.group(HC._help_node_first_pattern_value)
        
        if _key_second_node == None:
            _key_second_node = 1
            # print(_help_key+'_'+_key_first_node+'_'+str(_key_second_node))

        else:
            _key_second_node += 1
            _key_third_node = None
            _key_fourth_node = None
            # print(first_node_value)
        # 这里输出excel 子节点的节点 excel_key_child_node---> first_node_value
        excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR + \
            _key_first_node+STRING_CONNECTOR+str(_key_second_node)

        return excel_key_child_node,first_node_value
    elif not re.search(HC._help_node_first_pattern_1, line) == None:
            # pattern'    ▪︎︎请按照标准格式说出您想要采购的原料及其数量。'
        secondNodeMatch = re.search(HC._help_node_first_pattern_1, line)
        # print(secondNodeMatch)
        # second_node_value = secondNodeMatch.group(HC._help_node_first_pattern_value)
        first_node_value = secondNodeMatch.group(
            HC._help_node_first_pattern_value)

        if _key_second_node == None:
            _key_second_node = 1
            # print(_help_key+'_'+_key_first_node+'_'+str(_key_second_node))

        else:
            _key_second_node += 1
            _key_third_node = None
            _key_fourth_node = None
            # print(first_node_value)
        # 这里输出excel 子节点的节点 excel_key_child_node---> first_node_value
        excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR + \
            _key_first_node+STRING_CONNECTOR+str(_key_second_node)

        return excel_key_child_node, first_node_value
    elif not re.search(HC._help_node_first_pattern_2, line) == None:
            # pattern'    ▪︎︎请按照标准格式说出您想要采购的原料及其数量。'
        secondNodeMatch = re.search(HC._help_node_first_pattern_2, line)
        # print(secondNodeMatch)
        # second_node_value = secondNodeMatch.group(HC._help_node_first_pattern_value)
        first_node_value = secondNodeMatch.group(
            HC._help_node_first_pattern_value)

        if _key_second_node == None:
            _key_second_node = 1
            # print(_help_key+'_'+_key_first_node+'_'+str(_key_second_node))

        else:
            _key_second_node += 1
            _key_third_node = None
            _key_fourth_node = None
            # print(first_node_value)
        # 这里输出excel 子节点的节点 excel_key_child_node---> first_node_value
        excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR + \
            _key_first_node+STRING_CONNECTOR+str(_key_second_node)

        return excel_key_child_node, first_node_value
    elif not re.search(HC._help_node_second_pattern, line) == None:
        # pattern '        ▫︎如：胡萝卜采购100斤'
        nodeMatch = re.search(HC._help_node_second_pattern, line)
        second_node_value = nodeMatch.group(HC._help_node_second_pattern_value)
        # print('second_node_value'+second_node_value)
        if _key_third_node == None:
            _key_third_node = 'a'
        else:
            # a+1
            _key_third_node = chr(ord(_key_third_node)+1)
            _key_fourth_node = None
        # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value

        excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR+_key_first_node + \
            STRING_CONNECTOR+str(_key_second_node) + \
            STRING_CONNECTOR+_key_third_node
    
        return  excel_key_child_node,second_node_value

    elif not re.search(HC._help_node_third_pattern, line) == None:
        # pattern '            ৹【原料名称】为【胡萝卜】'
        nodeMatch = re.search(HC._help_node_third_pattern, line)
        third_node_value = nodeMatch.group(HC._help_node_third_pattern_value)

        # print('third_node_value'+third_node_value)
        if _key_fourth_node == None:
            _key_fourth_node = 1
        else:
            _key_fourth_node += 1
        # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value

        excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR+_key_first_node+STRING_CONNECTOR + \
            str(_key_second_node)+STRING_CONNECTOR+_key_third_node + \
            STRING_CONNECTOR+str(_key_fourth_node)


        return  excel_key_child_node,third_node_value

    elif not re.search(HC._help_node_foot_pattern, line) == None:
        nodeRootMatch = re.search(HC._help_node_foot_pattern, line)
        # pattern </StoreHouseListViewHelp>  != root
        value = nodeRootMatch.group(HC._help_node_foot_pattern_value)
        if not value == 'root':
            # 重置参数
            _help_key = None
            _help_key_version=None
            _init_key = False
            _key_first_node = None
            _key_second_node = None
            _special_node = None
            # print(value)
            return '',line
        else:
            # 文件读取结束
            return '',line
    else:
        # print(elif not re.search(r'(^\x00-\xff)!',line)==None:
        print('***************************匹配纯文字**************************'+line)
        line=re.sub(r'\n','',line)
        line=re.sub(r' *','',line)
        print('***************************匹配纯文字**************************'+line)

        if _special_node == None:
            _special_node = 1
        else:
            _special_node += 1
        if not _key_first_node==None:
            excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR + \
            SPECIAL_FLAG+STRING_CONNECTOR+str(_special_node)+STRING_CONNECTOR+'A_'
        else:
            # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value
            excel_key_child_node = _help_key+_help_key_version+STRING_CONNECTOR + \
            SPECIAL_FLAG+STRING_CONNECTOR+str(_special_node)
        return excel_key_child_node,line

def initExcel():
    wb = Workbook()
    return wb
def getSheetWs(wb,sheetTitle):
    ws = wb.worksheets[0]
    ws.title = sheetTitle
    ws.append({1: 'Key', 2: 'Value', 3: 'En', 4: 'Thai'})


    currentRow=ws._current_row
    cellHorizontal='center'

    cell=ws['A'+str(currentRow)]
    aCellColor='d3d3d3'
    cell.alignment=Alignment(horizontal=cellHorizontal,vertical=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=aCellColor,fgColor=aCellColor)

    cell=ws['B'+str(currentRow)] 
    bCellColor='fff0f5'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=bCellColor,fgColor=bCellColor)

    cell=ws['C'+str(currentRow)]
    cCellColor='ffe4e1'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=cCellColor,fgColor=cCellColor)

    cell=ws['D'+str(currentRow)]
    cell.alignment=Alignment(horizontal=cellHorizontal)
    dCellColor='ffc0cb'
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=dCellColor,fgColor=dCellColor)

    return ws

# 设置每个单元格的默认属性
def setCellProperty(worksheet):
    for i in range(1, worksheet.max_row+1):
        # worksheet.row_dimensions[i].height=50
        for col in range(ord('A'),ord('E')):
            worksheet[chr(col)+str(i)].font=Font(size=16)
            worksheet[chr(col)+str(i)].alignment=Alignment(wrap_text=True)
            borderStype='thick'
            borderColor='556b2f'
            worksheet[chr(col)+str(i)].border=Border(left=Side(border_style=borderStype,color=borderColor),
            right=Side(border_style=borderStype,color=borderColor),
            top=Side(border_style=borderStype,color=borderColor),
            bottom=Side(border_style=borderStype,color=borderColor))

    worksheet.column_dimensions['A'].width=45
    worksheet.column_dimensions['B'].width=85
    worksheet.column_dimensions['C'].width=90
    worksheet.column_dimensions['D'].width=90
    worksheet.column_dimensions['E'].width=90

def xmlConversionExcel(operatorType):
    wb = initExcel()
    global ws
    ws = getSheetWs(wb,sheet_title)
    # readFile(test_path, FILE_TYPE_ZH)
    
    readFile(help_zh_path, HC. _FILE_TYPE_ZH, operatorType)

    readFile(help_en_path, HC._FILE_TYPE_EN, operatorType)
    readFile(help_thai_path, HC. _FILE_TYPE_THAI, operatorType)
    
    errorFile = open(test_error_path, 'w+')
    for line in errorList:
        errorFile.write(line)
        errorFile.write('\n')
    errorFile.close()

    setCellProperty(ws)
    wb.save(su_string_file_path)

if __name__ == "__main__":
    xmlConversionExcel(_operator_type_excel)
    # xmlConversionExcel(_operator_type_key_xml)

