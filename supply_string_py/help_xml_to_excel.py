# 帮助文档转化为 excel


import re
import os
from openpyxl.workbook import Workbook
# 颜色
# 颜色
from openpyxl.styles import  PatternFill
from openpyxl.styles import Color, Fill,Font,colors,Border,Side,Alignment
from openpyxl.cell import Cell

import xlrd
from openpyxl.workbook import Workbook
import string_xml_to_excel as SE
import string_xml_constants as SC
from time import time

_help_key = None

_init_key = False
# A-Z
_key_first_node = None
# 1...
_key_second_node = None

# a-z
_key_third_node = None

# 1...
_key_fourth_node = None

# 1...
_special_node = None

# 特殊行进行特殊的标识
SPECIAL_FLAG = 'SP'

i = 1
j = 1
# A+1
# print(chr(ord(_key_first_node)+1))

INVALID_STRING_1 = '    <![CDATA['
INVALID_STRING_2 = '    ]]>'
STRING_CONNECTOR = '_'

PATTERN_STRING = r'    ▪︎︎'
FILE_TYPE_ZH = 'zh'
FILE_TYPE_EN = 'en'

# worksheets
ws=None
# 是否开启容错处理
isOpenFaultTolerantProcess=True
help_zh_path = '/Users/hehongqing/WorkSpace/Android/static_help/resources/xml/SupplyHelpFiles.xml'
help_en_path = '/Users/hehongqing/WorkSpace/Android/static_help/resources/xml/SupplyHelpENFiles.xml'
test_path = r'/Users/hehongqing/Downloads/test_help.xml'
su_string_file_path = r'/Users/hehongqing/Downloads/help_string_2.xlsx'
sheet_title = 'help'

def readFile(filePath, type):
    file = open(filePath)
    while 1:
        line = file.readline()
        if not line:
            file.close
            break
        else:
            # 读取到的每行
            # print(line)
            dealLine(line, type)

# <WarehouseManagerViewHelp version="V1" name="仓库管理">
#     <![CDATA[
# <b>添加仓库</b>
#     ▪︎︎点添加按钮，可添加新的仓库。

# <b>供货仓库</b>
#     ▪︎︎仓库可设置为供货仓库。
#     ▪︎︎设置为供货仓库的仓库，门店在创建采购单和退货单时，可指定向该仓库进行采购或退货。
#     ▪︎︎总部可设置多个供货仓库，也可不设置任何供货仓库。不设置供货仓库时，门店仍可向总部采购，只是不能指定向哪个仓库采购。
#     ]]>
# </WarehouseManagerViewHelp


def dealLine(line, type):
    global _init_key, _help_key, _key_first_node, _key_second_node, _key_third_node, _key_fourth_node, _special_node
    global i, j
    if _init_key == False and _help_key == None:
        match = re.search(r'version(.*)name', line)
        #  pattern  <WarehouseManagerViewHelp version="V1" name="仓库管理">
        if not match == None:
            _init_key = True

            # print(match)
            nodeTitleMatch = re.search(
                r'<(?P<Key>\w+)(.*)version=\"(?P<VersionName>\w[0-9])"(.*)name=\"(?P<Title>.*)\"', line)
            print(nodeTitleMatch)
            print('key          '+nodeTitleMatch.group('Key'))
            _help_key = nodeTitleMatch.group('Key')
            nodeVersion = nodeTitleMatch.group('VersionName')
            print('VersionName  '+nodeVersion)

            nodeTitleValue = nodeTitleMatch.group('Title')
            print('Title        '+nodeTitleValue)
            excel_key_title = _help_key+STRING_CONNECTOR+nodeVersion
            if type == FILE_TYPE_ZH:
               # 在这里输出title到excel中  excel_key_title--->nodeTitleValue
                ws.append({i: excel_key_title, 2: nodeTitleValue})
            elif type == FILE_TYPE_EN:
                j +=1
                # 进行判断
                cell=ws['A'+str(j)]
                # 如果这里的key无法进行匹配 直接标红 需要检查xml文件
                if not cell._value ==excel_key_title and isOpenFaultTolerantProcess:
                    cell=ws['E'+str(j)]
                    enEmptyColor='dc143c'
                    cell.fill=PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
                    ws['E'+str(j)]=excel_key_title

                ws['A'+str(j)]=excel_key_title
                cell=ws['B'+str(j)]
                # 英文和中文一致，则认为没有进行翻译，输出值为空，并且做出标记
                if cell._value==nodeTitleValue:
                    ws['C'+str(j)]=''
                    enEmptyColor='dda0dd'
                    ws['C'+str(j)].fill==PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
                else:
                    ws['C'+str(j)]=nodeTitleValue
            return
            # excel 输出结构
            #  key                              value                               en
            #  key_version                      title
            #  WarehouseManagerViewHelp_V1      仓库管理
            #  key_A_0                          添加仓库（截取掉<b>*</b>）
            #  key_A_1                          点添加按钮，可添加新的仓库。(截取掉    ▪︎︎)
            #  key_B_0                          供货仓库（截取掉<b>*</b>）
            #  key_B_1                          仓库可设置为供货仓库。(截取掉    ▪︎︎)
            #  ...依次递加
    elif INVALID_STRING_1 in line or INVALID_STRING_2 in line or line == '\n' or not re.search(r'^ *\n', line) == None or '<![CDATA[' in line or not re.search(r'	]]>',line)==None or ']]>'in line:
        # print('无效返回')
        return
    elif not re.search(r'<b>(.*)</b>', line) == None:
        # pattern <b>添加仓库</b>
        nodeMatch = re.search(r'<b>(.*)</b>', line)
        first_node_value = nodeMatch.group(1)
        if _key_first_node == None:
            _key_first_node = 'A'
            print(first_node_value)
            # 在这里输出 excel 第一个子节点  excel_key_node--->first_node_value
            excel_key_node = _help_key+STRING_CONNECTOR+_key_first_node
            if type==FILE_TYPE_EN:
                j+=1
                ws['A'+str(j)]=excel_key_node
                ws['C'+str(j)]=first_node_value
            else:
                ws.append({i: excel_key_node, 2: first_node_value})
            return
        else:
            # A+1
            # 同时需要将节点 置为初始值
            _key_second_node = None
            _key_third_node = None
            _key_fourth_node = None

            _key_first_node = chr(ord(_key_first_node)+1)
            # 在这里输出 excel 第一个子节点  excel_key_node--->first_node_value
            excel_key_node = _help_key+STRING_CONNECTOR+_key_first_node

            if type==FILE_TYPE_EN:
                j +=1
                ws['A'+str(j)]=excel_key_node
                ws['C'+str(j)]=first_node_value
            else:
                ws.append({i: excel_key_node, 2: first_node_value})
            return

    elif not re.search(r'    ▪︎︎(.*)', line) == None:
        # pattern'    ▪︎︎请按照标准格式说出您想要采购的原料及其数量。'
        secondNodeMatch = re.search(r'    ▪︎︎(.*)', line)
        print(secondNodeMatch)
        second_node_value = secondNodeMatch.group(1)
        if _key_second_node == None:
            _key_second_node = 1
            print(_help_key+'_'+_key_first_node+'_'+str(_key_second_node))
            # excel_key_child_node = _help_key+STRING_CONNECTOR + \
            #     _key_first_node+STRING_CONNECTOR+str(_key_second_node)
            # ws.append({i: excel_key_child_node, j: second_node_value})
        else:
            _key_second_node += 1
            _key_third_node = None
            _key_fourth_node = None
            print(second_node_value)
        # 这里输出excel 子节点的节点 excel_key_child_node---> second_node_value
        excel_key_child_node = _help_key+STRING_CONNECTOR + \
            _key_first_node+STRING_CONNECTOR+str(_key_second_node)

        if type==FILE_TYPE_EN:
            j+=1
            ws['A'+str(j)]=excel_key_child_node
            ws['C'+str(j)]=second_node_value
        else:
            ws.append({i: excel_key_child_node, 2: second_node_value})
        return
    elif not re.search(r'        ▫︎(.*)', line) == None:
        # pattern '        ▫︎如：胡萝卜采购100斤'
        thirdNodeMatch = re.search(r'        ▫︎(.*)', line)
        print(thirdNodeMatch)
        third_node_value = thirdNodeMatch.group(1)
        if _key_third_node == None:
            _key_third_node = 'a'
        else:
            # a+1
            _key_third_node = chr(ord(_key_third_node)+1)
            _key_fourth_node = None
        # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value

        excel_key_child_node = _help_key+STRING_CONNECTOR+_key_first_node + \
            STRING_CONNECTOR+str(_key_second_node) + \
            STRING_CONNECTOR+_key_third_node
        if type==type == FILE_TYPE_EN:
            j +=1
            ws['A'+str(j)]=excel_key_child_node
            ws['C'+str(j)]=third_node_value
        else:
            ws.append({i: excel_key_child_node, 2: third_node_value})
        return
    elif not re.search(r'            ৹(.*)', line) == None:
        # pattern '            ৹【原料名称】为【胡萝卜】'
        fourthNodeMatch = re.search(r'            ৹(.*)', line)
        print(fourthNodeMatch)
        fourth_node_value = fourthNodeMatch.group(1)
        if _key_fourth_node == None:
            _key_fourth_node = 1
        else:
            _key_fourth_node += 1
        # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value

        excel_key_child_node = _help_key+STRING_CONNECTOR+_key_first_node+STRING_CONNECTOR + \
            str(_key_second_node)+STRING_CONNECTOR+_key_third_node + \
            STRING_CONNECTOR+str(_key_fourth_node)
        if type== FILE_TYPE_EN:
            j +=1
            ws['A'+str(j)]=excel_key_child_node
            ws['C'+str(j)]=fourth_node_value
        else:
            ws.append({i: excel_key_child_node, 2: fourth_node_value})
        return
    elif not re.search(r'^\</(.*)>', line) == None:
        nodeRootMatch = re.search(r'^</(.*)>', line)
        # pattern </StoreHouseListViewHelp>  != root
        value = nodeRootMatch.group(1)
        if not value == 'root':
            # 重置参数
            _help_key = None
            _init_key = False
            _key_first_node = None
            _key_second_node = None
            _special_node = None
            print(value)
            return
        else:
            # 文件读取结束
            return
    else:
        # print(elif not re.search(r'(^\x00-\xff)!',line)==None:
        print('***************************匹配纯文字**************************'+line)
        line=re.sub(r'\n','',line)
        line=re.sub(r' *','',line)
        print('***************************匹配纯文字**************************'+line)

        if _special_node == None:
            _special_node = 1
        else:
            _special_node += 1
        if not _key_first_node==None:
            excel_key_child_node = _help_key+STRING_CONNECTOR + \
            SPECIAL_FLAG+STRING_CONNECTOR+str(_special_node)+STRING_CONNECTOR+'A_'
        else:
            # 这里输出excel 子节点的节点 excel_key_child_node---> fourth_node_value
            excel_key_child_node = _help_key+STRING_CONNECTOR + \
            SPECIAL_FLAG+STRING_CONNECTOR+str(_special_node)

        if type== FILE_TYPE_EN:
            j +=1
            ws['A'+str(j)]=excel_key_child_node
            ws['C'+str(j)]=line
        else:
            ws.append({i: excel_key_child_node, 2: line})
        return


def getSheetWs(wb,sheetTitle):
    ws = wb.worksheets[0]
    ws.title = sheetTitle
    ws.append({1: 'Key', 2: 'Value', 3: 'En', 4: 'Thai'})


    currentRow=ws._current_row
    cellHorizontal='center'

    cell=ws['A'+str(currentRow)]
    aCellColor='d3d3d3'
    cell.alignment=Alignment(horizontal=cellHorizontal,vertical=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=aCellColor,fgColor=aCellColor)

    cell=ws['B'+str(currentRow)] 
    bCellColor='fff0f5'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=bCellColor,fgColor=bCellColor)

    cell=ws['C'+str(currentRow)]
    cCellColor='ffe4e1'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=cCellColor,fgColor=cCellColor)

    cell=ws['D'+str(currentRow)]
    cell.alignment=Alignment(horizontal=cellHorizontal)
    dCellColor='ffc0cb'
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=dCellColor,fgColor=dCellColor)

    return ws


def initExcel():
    wb = Workbook()
    return wb

# 设置每个单元格的默认属性
def setCellProperty(worksheet):
    for i in range(1, worksheet.max_row+1):
        # worksheet.row_dimensions[i].height=50
        for col in range(ord('A'),ord('D')):
            worksheet[chr(col)+str(i)].font=Font(size=16)
            worksheet[chr(col)+str(i)].alignment=Alignment(wrap_text=True)
            borderStype='thick'
            borderColor='556b2f'
            worksheet[chr(col)+str(i)].border=Border(left=Side(border_style=borderStype,color=borderColor),
            right=Side(border_style=borderStype,color=borderColor),
            top=Side(border_style=borderStype,color=borderColor),
            bottom=Side(border_style=borderStype,color=borderColor))

    worksheet.column_dimensions['A'].width=60
    worksheet.column_dimensions['B'].width=90
    worksheet.column_dimensions['C'].width=90
    worksheet.column_dimensions['D'].width=90

def xmlConversionExcel():
    wb = initExcel()
    global ws
    ws = getSheetWs(wb,sheet_title)
    # readFile(test_path, FILE_TYPE_ZH)

    readFile(help_zh_path,FILE_TYPE_ZH)

    readFile(help_en_path,FILE_TYPE_EN)

    setCellProperty(ws)

    wb.save(su_string_file_path)

if __name__ == "__main__":
    xmlConversionExcel()

