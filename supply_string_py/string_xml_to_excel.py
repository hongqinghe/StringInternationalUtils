#  xml  文件转化为 excel 文件

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import os
import string
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element, SubElement, ElementTree
from openpyxl.workbook import Workbook
# 颜色
from openpyxl.styles import  PatternFill
from openpyxl.styles import Color, Fill,Font,colors,Border,Side,Alignment
from openpyxl.cell import Cell

import string_xml_constants as SC
# import xlwt

# su_string_file_path=r'/Users/hehongqing/Downloads/supply_string_4.xlsx'
su_string_file_path=r'/Users/hehongqing/Downloads/supply_string_3.xlsx'

# 获取每个sheet
def getSheetWs(wb,sheet_title):
    if SC._supply_base_sheet_head_btn==sheet_title:
        ws = wb.worksheets[0]
        ws.title = sheet_title
    else:
        ws = wb.create_sheet(sheet_title)
    ws.append({1: SC._excel_title_value_key, 2:SC._excel_title_value_value, 3: SC._excel_title_value_en, 4: SC._excel_title_value_thai})

    currentRow=ws._current_row
    cellHorizontal='center'

    cell=ws['A'+str(currentRow)]
    aCellColor='d3d3d3'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=aCellColor,fgColor=aCellColor)

    cell=ws['B'+str(currentRow)] 
    bCellColor='fff0f5'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=bCellColor,fgColor=bCellColor)

    cell=ws['C'+str(currentRow)]
    cCellColor='ffe4e1'
    cell.alignment=Alignment(horizontal=cellHorizontal)
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=cCellColor,fgColor=cCellColor)

    cell=ws['D'+str(currentRow)]
    cell.alignment=Alignment(horizontal=cellHorizontal)
    dCellColor='ffc0cb'
    cell.fill=PatternFill(fill_type='lightHorizontal',bgColor=dCellColor,fgColor=dCellColor)
    return ws

def initExcel(path):
    wb = Workbook()
    fileName = path
    # 开始写入 excel
    return  wb, fileName
 
# ws= getOnlyWs(TYPE_SUPPLY_BASE_BTNS)
def xmlParse(zhPath, enPath,thaiPath,ws):
    zh_tree = ET.parse(zhPath)
    en_tree = ET.parse(enPath)
    thai_tree = ET.parse(thaiPath)
    # 中文 xml  解析 生成对应的 key 和 value 
    for zh_elem in zh_tree.iter(tag='string'):
        zh_key = zh_elem.attrib['name']
        print('zh_key:      ' + str(zh_key))
        zh_value = zh_elem.text
        print('zh_value:    ' + str(zh_value))

        # 获取 key 对应的 en_value
        en_value=''
        for en_elem in en_tree.iter(tag='string'):
            en_key = en_elem.attrib['name']
            if en_key == zh_key:
                en_value = en_elem.text
                break

        # 获取 key 对应的 thai_value
        thai_value = ''
        for thai_elem in thai_tree.iter(tag='string'):
            thai_key = thai_elem.attrib['name']
            if thai_key == zh_key:
                thai_value = thai_elem.text
                break
        #  添加字符串到excel中
        ws.append([zh_key, zh_value, en_value, thai_value])
        # 如果en_value 字符为空，则认为该项没有翻译，进行颜色填充
        if en_value=='':
            currentRow=ws._current_row
            cell=ws['C'+str(currentRow)]
            enEmptyColor='dda0dd'
            cell.fill=PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)
        # if thai_value == '' or  re.findall('%s',zh_value)==1:
        if not zh_value == None:
            if (zh_value.count('%s') > 0 and thai_value.count('%s') <= 0) or (zh_value.count('%d') > 0 and thai_value.count('%d') <= 0) or (zh_value.count('%1$s') > 0 and thai_value.count('%1$s') <= 0) or (thai_value == '') or (zh_value.count('\n')>0 and thai_value.count('\n')<=0):
                currentRow = ws._current_row
                cell = ws['D'+str(currentRow)]
                thaiEmptyColor = 'dda033'
                cell.fill = PatternFill(
                    fill_type='lightGray', bgColor=thaiEmptyColor, fgColor=thaiEmptyColor)
            # cell=ws['D'+str(currentRow+1)]
            # enEmptyColor='ff69b4'
            # cell.fill=PatternFill(fill_type='lightGray',bgColor=enEmptyColor,fgColor=enEmptyColor)

# 设置每个单元格的默认属性
def setCellProperty(worksheet):
    for i in range(1, worksheet.max_row+1):
        worksheet.row_dimensions[i].height=50
        for col in range(ord('A'),ord('E')):
            worksheet[chr(col)+str(i)].font=Font(size=14)
            worksheet[chr(col)+str(i)].alignment=Alignment(wrap_text=True)
            borderStype='thick'
            borderColor='556b2f'
            worksheet[chr(col)+str(i)].border=Border(left=Side(border_style=borderStype,color=borderColor),
            right=Side(border_style=borderStype,color=borderColor),
            top=Side(border_style=borderStype,color=borderColor),
            bottom=Side(border_style=borderStype,color=borderColor))

    worksheet.column_dimensions['A'].width=45
    worksheet.column_dimensions['B'].width=90
    worksheet.column_dimensions['C'].width=115
    worksheet.column_dimensions['D'].width=115

# 转义为excel
def transformToExcel():

    # 初始化excel 的必要参数
    wb, fileName = initExcel(su_string_file_path)
    ws=getSheetWs(wb,SC._supply_base_sheet_head_btn)
    # supplyBase
    xmlParse(SC._supplybase_zh_btns_path, SC._supplybase_en_btns_path,
             SC._supplybase_thai_btns_path,ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._supply_base_sheet_head_page)
    xmlParse(SC._supplybase_zh_pages_path,
             SC._supplybase_en_pages_path, SC._supplybase_thai_pages_path, ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._supply_base_sheet_head_msg)
    xmlParse(SC._supplybase_zh_msgs_path, SC._supplybase_en_msgs_path,
             SC._supplybase_thai_msgs_path, ws)
    setCellProperty(ws)

    # base
    ws=getSheetWs(wb,SC._base_sheet_head_btn)
    xmlParse(SC._base_zh_btns_path, SC._base_en_btns_path,
             SC._base_thai_btns_path, ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._base_sheet_head_page)
    xmlParse(SC._base_zh_pages_path, SC._base_en_pages_path,
             SC._base_thai_pages_path, ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._base_sheet_head_msg)
    xmlParse(SC._base_zh_msgs_path, SC._base_en_msgs_path,
             SC._base_thai_msgs_path, ws)
    setCellProperty(ws)

    # Buy
    ws=getSheetWs(wb,SC._buy_sheet_head_btn)
    xmlParse(SC._buy_zh_btns_path, SC._buy_en_btns_path,
             SC._buy_thai_btns_path, ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._buy_sheet_head_page)
    xmlParse(SC._buy_zh_pages_path, SC._buy_en_pages_path,
             SC._buy_thai_pages_path, ws)
    setCellProperty(ws)

    ws=getSheetWs(wb,SC._buy_sheet_head_msg)
    xmlParse(SC._buy_zh_msgs_path, SC._buy_en_msgs_path,
             SC._buy_thai_msgs_path, ws)
    setCellProperty(ws)

    # 在每个 ws 对象添加节点后对文件做保存操作，不然不会生效
    wb.save(fileName)
if __name__ == "__main__":
    transformToExcel()
